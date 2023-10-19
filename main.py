import openpyxl
import xml.etree.ElementTree as ET
from pathlib import Path
from decimal import Decimal
import pyodbc
from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter import messagebox as msb
import os
import fileinput


# deklaracja lokalizacji plików i zmiennych

dokument_kalkulacji = 'N:/Wsp-Tym/ESM/autoWycena/nr_wyceny.cprj'
dokument_kalkulacji_wynik = 'N:/Wsp-Tym/ESM/autoWycena/nr_wyceny_out.cprj'
dokument_TKW = 'N:\\Wsp-Tym\\ESM\\autoWycena\\TKW-testowe.xlsm'

test_path = "N:/Wsp-Tym/ESM/autoWycena/nr_wyceny_txt.txt"

blacha = {"inox": "1.4301", "czarna": "St37", "aluminium": "AlMg3"}

lista_kodow = []

#########################################################################################
######## FUNKCJE #################################


def check_next_row(n, ws4):
    for _ in range(3):
        n += 1
        if ws4.cell(row=n, column=2).value != None:
            return n
    return 0

    # sprawdza czy kolejne 3 wiersze sa puste jezeli tak zwraca wartosc empty_check  = True
    # jezeli nie zwraca numer wiersza w ktorym trzeba wznowic szukanie danych


def read_TKW():
    """## Pobranie z pliku TKW z arkusza "wycena":
    ### Kod
    ### Rysunek
    ### ilosc na dostawę
    ### Rodaj materiału
    ### (grubość blachy)
    """

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(
        dokument_TKW, read_only=True, data_only=True)
    ws4 = wb_obj["wycena"]

    empty_check = False
    n = 19

    # petla pobierajaca listę kodow z dokumentu TKW
    while not empty_check:

        if ws4.cell(row=n, column=20).value:
            KOD = ws4.cell(row=n, column=2)
            Rysunek = ws4.cell(row=n, column=4)
            ilosc_na_dostawe = ws4.cell(row=n, column=7)
            rodzaj_materialu = ws4.cell(row=n, column=20)
            grubosc_blachy = ws4.cell(row=n, column=31)

            # błąd niekompletne dane:
            if grubosc_blachy.value is None:
                msb.showinfo(
                    title="Błąd", message="Niekompletne dane o grubosci blachy dla kodu: " + str(KOD.value) + "\nProgram zostanie zamknięty, popraw dane.")
                exit()
            if ilosc_na_dostawe.value is None:
                msb.showinfo(
                    title="Błąd", message="Niekompletne dane o ilości na dostawę dla kodu: " + str(KOD.value) + "\nProgram zostanie zamknięty, popraw dane.")
                exit()
                kod = ''
                rysunek = ''
            if KOD.value is not None:
                kod = str(KOD.value).upper()
            else:
                kod = None
            if Rysunek.value is not None:
                rysunek = str(Rysunek.value).upper()
            else:
                rysunek = None

            kod = {"kod": kod, "rysunek": str(rysunek),
                   "ilosc_na_dostawe": ilosc_na_dostawe.value, "rodzaj_materialu": rodzaj_materialu.value, "grubosc_blachy": grubosc_blachy.value}

            print(kod)
            lista_kodow.append(kod)

        n += 1

        if ws4.cell(row=n, column=2).value == None:
            n = check_next_row(n, ws4)
            if n == 0:
                empty_check = True


def remove_exponent(d):
    """usuwa 0 po przecinku"""
    return d.quantize(Decimal(1)) if d == d.to_integral() else d.normalize()


def get_m_id(m_name):
    """Get m_id from database basis on m_name"""
    cnxn = pyodbc.connect(
        'DRIVER={SQL Server};SERVER=TRUMPF\TRUMPFSQL2;DATABASE=TCALCUSER_V01_V180000;UID=trudbuser;PWD=foll_subbr')
    cursor = cnxn.cursor()

    cursor.execute(
        "SELECT m_id FROM [TCALCUSER_V01_V180000].[dbo].[TcMaterial] WHERE m_name ='"+m_name+"'")
    for row in cursor.fetchall():
        return row.m_id


def BasicMaterialInformation(rodzaj_materialu):
    """Get m_id from database basis on m_name"""
    cnxn = pyodbc.connect(
        'DRIVER={SQL Server};SERVER=TRUMPF\TRUMPFSQL2;DATABASE=TCALCUSER_V01_V180000;UID=trudbuser;PWD=foll_subbr')
    cursor = cnxn.cursor()

    cursor.execute(
        "SELECT * FROM [TCALCUSER_V01_V180000].[dbo].[TcWerkstoff] WHERE m_name ='"+rodzaj_materialu+"'")
    for row in cursor.fetchall():
        # print(row)
        return row


def get_value_from_dict(dic_element, key):
    value = dic_element.get(key)
    value = value or "PUSTY!@#"
    return value


def m_name(rodzaj_materialu, grubosc_blachy):
    """konwertuje do formatu XX"""
    kod_materialu = blacha.get(rodzaj_materialu)
    grubosc_kod = str(grubosc_blachy*10)

    # tes = len(grubosc_kod.split('.')[0])
    grubosc_kod = str(remove_exponent(Decimal(grubosc_kod)))
    if len(grubosc_kod.split('.')[0]) < 2:
        grubosc_kod = '0'+grubosc_kod

    return kod_materialu + '-' + grubosc_kod


def Resource_articleNO(rodzaj_materialu, grubosc_blachy):
    kod_materialu = blacha.get(rodzaj_materialu)
    grubosc_kod = str(grubosc_blachy)
    if len(grubosc_kod) < 2:
        grubosc_kod = '0'+grubosc_kod

    # print(kod_materialu + '-' + grubosc_kod+'x3000x1500')
    return kod_materialu + '-' + grubosc_kod+'x3000x1500'


def zamien_znaki(filename):
    with open(filename, 'r', encoding='utf-8') as file:
        content = file.read()

    content = content.replace('ł', 'l').replace('Ł', 'L')
    content = content.replace('ń', 'n').replace('Ń', 'N')
    content = content.replace('ż', 'z').replace('Ż', 'Z')
    content = content.replace('ó', 'o').replace('Ó', 'O')
    content = content.replace('ź', 'z').replace('Ź', 'Z')

    with open(filename, 'w', encoding='utf-8') as file:
        file.write(content)


def parse_xml(path):
    # zamiana polskich znaków
    zamien_znaki(path)

    tree = ET.parse(path)
    root = tree.getroot()
    # parsing using the string.
    # stringroot = ET.fromstring(data)
    # printing the root.
    for neighbor in root.iter('document'):
        # print(neighbor.attrib)
        # finding the state tag and their child attributes.
        for state in root.findall('body'):
            for el in state.findall('Parts'):
                for el2 in el.findall('Part'):
                    #########
                    # test
                    # for dfggfd in el2.findall('Material'):
                    #     for elem in dfggfd.iter():
                    #         print(elem.tag)

                    # os.system('cls||clear')
                    # end test
                    # for elem in el2.iter():
                    #     print(elem.tag)
                    # os.system('cls||clear')
                    #######
                    # print(el2.find('ArticleNo').text)
                    # print(el2.find('ArticleNo').text)
                    kod = el2.find('ArticleNo').text.upper()
                    for n in lista_kodow:
                        kod_check = get_value_from_dict(n, "kod")
                        rysunek_check = get_value_from_dict(n, "rysunek")
                        if kod_check in kod or rysunek_check in kod:

                            rodzaj_materialu = n.get(
                                "rodzaj_materialu")
                            grubosc_blachy = n.get(
                                "grubosc_blachy")
                            # print(n.get("rysunek"))
                            # ilosc = n.get("ilosc_na_dostawe")
                            # print(str("%.6f" % n.get("ilosc_na_dostawe")))
                            # print(el2.find('RawMaterialName').text)

                            # zmiana nazwy materialu //PART/RawMaterialName
                            for el3 in el2.findall('RawMaterialName'):

                                # pobranie nazwy_materaiłu zgodnie z kodowaniem calculate
                                kod_materialu = m_name(
                                    n.get("rodzaj_materialu"), n.get("grubosc_blachy"))
                                el3.text = kod_materialu

                            # zmiana id materiału //PART/RawMaterialName
                            for el4 in el2.findall('RawMaterialID'):
                                # print(el2.find('RawMaterialID').text)
                                # dfdg = get_m_id(kod_materialu)
                                # print(type(dfdg))
                                id_materialu = str(get_m_id(kod_materialu))
                                el4.text = id_materialu

                            # usuniecie tagu material in Part
                            # test
                            # for elnnn1 in el2.findall('.//Material'):
                            #     for elnnn2 in elnnn1.findall('MaterialId'):
                            #         elnnn2.text = id_materialu
                            #     for elnnn2 in elnnn1.findall('MaterialName'):
                            #         elnnn2.text = kod_materialu
                            # edn test
                            for elnnn1 in el2.findall('Material'):
                                for elnnn2 in elnnn1.findall('MaterialId'):
                                    elnnn2.text = id_materialu
                                for elnnn2 in elnnn1.findall('MaterialName'):
                                    elnnn2.text = kod_materialu

                                # dodanie tagu Material z dwoma tagami: MaterialId i MaterialName

                                # ET.SubElement(el2, 'Material')
                                # for temp in el2.findall('Material'):
                                #     ET.SubElement(temp, 'MaterialId')
                                #     ET.SubElement(temp, 'MaterialName')
                                #     for gfg in temp.findall('MaterialId'):
                                #         gfg.text = id_materialu
                                #     for gfg in temp.findall('MaterialName'):
                                #         gfg.text = kod_materialu

                                # for temp in el2.iter('Material'):
                                    # print(temp.tag, temp.text)
                                    # if temp.tag == 'MaterialName':
                                    #     temp.text = kod_materialu
                                    # elif temp.tag == 'MaterialId':
                                    #     temp.text = id_materialu
                                    # elif temp.tag == 'Material':
                                    #     sgdg = 1

                            # for eln1 in el2.findall("Material"):
                            #     eln1.remove("Material")

                            # zmiana materialu w /Part/WorkingPlan/WorkingSteps/WorkingStep/Resources/Resource

                            for elnn1 in el2.findall('WorkingPlan'):
                                for elnn2 in elnn1.findall('WorkingSteps'):
                                    for elnn3 in elnn2.findall('WorkingStep'):
                                        for elnn4 in elnn3.findall('Resources'):
                                            for elnn5 in elnn4.findall('Resource'):
                                                # print(elnn5.find('ArticleNo').text)
                                                for elem in elnn5.iter():
                                                    # print(elem.tag, elem.text)
                                                    if elem.tag == 'RawMaterialID':
                                                        elem.text = id_materialu
                                                    elif elem.tag == 'RawMaterialName':
                                                        elem.text = kod_materialu
                                                    elif elem.tag == 'ArticleNo':
                                                        elem.text = Resource_articleNO(
                                                            rodzaj_materialu, grubosc_blachy)
                                                        # elif elem.tag == 'Resource':
                                                        #     ssdfsd = 1
                                                        # else:
                                                        #     elnn5.remove(elem)

                            # zmiana materialu w /Part/Material/BasicMaterial
                            basic_material_information = BasicMaterialInformation(
                                blacha.get(rodzaj_materialu))

                            for elnnn1 in el2.findall('Material'):
                                for elnnn2 in elnnn1.findall('BasicMaterial'):
                                    for elem in elnnn2.iter():
                                        # print(elem.tag, elem.text)
                                        if elem.tag == 'BasicMaterialName':
                                            if basic_material_information.m_name == "St37":
                                                elem.text = "1.0038"
                                            else:
                                                elem.text = basic_material_information.m_name
                                        elif elem.tag == 'BasicMaterialId':
                                            elem.text = str(
                                                basic_material_information.m_id)
                                        elif elem.tag == 'BasicMaterialDescription':
                                            elem.text = basic_material_information.m_bezeichnung
                                        elif elem.tag == 'BasicMaterialIsSystem':
                                            elem.text = str(
                                                basic_material_information.m_id)
                                        elif elem.tag == 'TensileStrength':
                                            elem.text = str(
                                                basic_material_information.m_zugFestigkeit_m)
                                        elif elem.tag == 'SpecificWeight':
                                            elem.text = str(
                                                basic_material_information.m_spezGewicht_m)

         # zmiana ilosci w Part ID=1

                    for eln1 in el2.findall('SubItems'):
                        for eln2 in eln1.findall('SubItem'):
                            kod = eln2.find('ArticleNo').text.upper()
                            for n in lista_kodow:
                                kod_check = get_value_from_dict(
                                    n, "kod")
                                rysunek_check = get_value_from_dict(
                                    n, "rysunek")
                                if kod_check in kod or rysunek_check in kod:
                                    for el in eln2.findall('Quantity'):
                                        # elof=float(.get("ilosc_na_dostawe"))
                                        el.text = str("%.6f" %
                                                      n.get("ilosc_na_dostawe"))

                                        # print(str("%.6f" %
                                        #       n.get("ilosc_na_dostawe")))
                                        # el.text = n.get("ilosc_na_dostawe")
                                        # print(str("%.6f" %
                                        #       n.get("ilosc_na_dostawe")))

    global dokument_kalkulacji_wynik
    dokument_kalkulacji_wynik = dokument_kalkulacji[:-5]+"_POP.cprj"
    tree.write(dokument_kalkulacji_wynik)
    # name = state.get('ArticleNo')
    # print(name, rank)

    # print(root)
    # print(stringroot)


##################################################
# def test(path):
#     # importing the module.
#     import xml.etree.ElementTree as ET
#     XMLexample_stored_in_a_string = '''<?xml version ="1.0"?>
#     <COUNTRIES>
#         <country name ="INDIA">
#             <neighbor name ="Dubai" direction ="W"/>
#         </country>
#         <country name ="Singapore">
#             <neighbor name ="Malaysia" direction ="N"/>
#         </country>
#     </COUNTRIES>
#     '''
#     # parsing directly.
#     tree = ET.parse(path)
#     root = tree.getroot()
#     # parsing using the string.
#     stringroot = ET.fromstring(XMLexample_stored_in_a_string)
#     # printing the root.
#     print(root)
#     print(stringroot)


####################################################
# def test2():
#     import xml.etree.ElementTree as ET
#     XMLexample_stored_in_a_string = '''<?xml version ="1.0"?>
#     <States>
#         <state name ="TELANGANA">
#             <rank>1</rank>
#             <neighbor name ="ANDHRA" language ="Telugu"/>
#             <neighbor name ="KARNATAKA" language ="Kannada"/>
#         </state>
#         <state name ="GUJARAT">
#             <rank>2</rank>
#             <neighbor name ="RAJASTHAN" direction ="N"/>
#             <neighbor name ="MADHYA PRADESH" direction ="E"/>
#         </state>
#         <state name ="KERALA">
#             <rank>3</rank>
#             <neighbor name ="TAMILNADU" direction ="S" language ="Tamil"/>
#         </state>
#     </States>
#     '''
#     # parsing from the string.
#     root = ET.fromstring(XMLexample_stored_in_a_string)
#     # printing attributes of the root tags 'neighbor'.
#     for neighbor in root.iter('neighbor'):
#         print(neighbor.attrib)
#     # finding the state tag and their child attributes.
#     for state in root.findall('state'):
#         rank = state.find('rank').text
#         name = state.get('name')
#         print(name, rank)

###################################################
# MAIN
# Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
# show an "Open" dialog box and return the path to the selected file

msb.showinfo(title="Wybierz plik calculate", message="Wybierz plik calculate")
dokument_kalkulacji = askopenfilename()

msb.showinfo(title="Wybierz plik TKW", message="Wybierz plik TKW")
dokument_TKW = askopenfilename()
# odczytanie danych z pliku TKW
read_TKW()


# parsownie pliku XML wyceny_calculate
parse_xml(dokument_kalkulacji)
# test(dokument_kalkulacji)
# test2()
print("Koniec programu")
msb.showinfo(title="KONIEC", message="poprawiony plik: " +
             dokument_kalkulacji_wynik)
